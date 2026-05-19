import openpyxl
from datetime import datetime
import json, os

EXCEL = r'C:\Users\Administrator\Desktop\Dashboard\backend\Files\Economic Quarterly data Final.xlsx'
OUT = r'C:\Users\Administrator\Desktop\Dashboard\frontend\src\data\industries.js'

wb = openpyxl.load_workbook(EXCEL)

def extract_series(ws):
    start_row_idx = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and len(row) > 0 and row[0] == 'Start':
            start_row_idx = idx
            break
    if start_row_idx is None:
        return {}

    header_row = list(ws.iter_rows(min_row=start_row_idx, max_row=start_row_idx, values_only=True))[0]
    series_cols = []
    for col in range(3, len(header_row)):
        if header_row[col] is not None:
            series_cols.append((str(header_row[col]).strip(), col, col + 1))

    results = {}
    data_start_row = start_row_idx + 2
    for name, dc, vc in series_cols:
        data = []
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if row[dc] is not None and row[vc] is not None:
                try:
                    d = row[dc]; v = row[vc]
                    if isinstance(d, datetime) and isinstance(v, (int, float)):
                        data.append([d.strftime('%b %Y'), round(float(v), 4)])
                except: pass
        results[name] = data
    return results

A = {}
for sheet in wb.sheetnames:
    A[sheet] = extract_series(wb[sheet])

def lbl(s): return [x[0] for x in s]
def vals(s): return [x[1] for x in s]

def kch(s, is_abs=False, suffix='', rnd=2):
    if len(s) < 2: return '+0', True
    c = s[-1][1]; p = s[-2][1]; ch = c - p; up = ch >= 0; sg = '+' if up else ''
    if is_abs: return f'{sg}{round(ch, rnd)}{suffix}', up
    if abs(p) > 0:
        pch = (ch/abs(p))*100; return f'{sg}{round(pch,1)}%', up
    return '+0%', True

QL = ['Mar 2025','Jun 2025','Sep 2025','Dec 2025','Feb 2026']
QC = ['Indicator','Mar 2025','Jun 2025','Sep 2025','Dec 2025','Feb 2026','Trend']
KS = ['q3_23','q4_23','q1_24','q2_24','q3_24']

def tr(s): return '\u25b2' if len(s)<2 or s[-1][1]>=s[-2][1] else '\u25bc'

def qrow(name, s, fmt=None):
    d = dict(s); row = {'indicator': name}
    for ki, m in zip(KS, QL):
        v = d.get(m)
        row[ki] = fmt(v) if (v is not None and fmt) else (str(round(v,2)) if v is not None else '-')
    row['trend'] = tr(s)
    return row

def cp(name):
    raw = A['Commodity Price ']
    for k in raw:
        if k.strip() == name: return raw[k]
    raise KeyError(name)

ei=A['Economic Indicators']; im=A['Industrial and Manufacturing']
tr2=A['Transportation and Automotive']; se=A['Services and exports']
bf=A['Banking and Finance']; ty=A['Treasury Yield']; idx=A['Index']

gst=ei['GST Collections']; cpi=ei['Retail Inflation (CPI)']
trade=ei['India Trade Balance']; pmi_c=ei['Composite PMI']
vix_i=ei['India VIX']; rate_i=ei['India Interest Rate']
usd_inr=ei['Dollar Rupee Exchange Rate']; dxy=ei['DXY Index']
gold_res=ei['India Gold Reserve']
fx_res=ei['India Total FX Reserves']
fx_res_usd=ei['India FX Reserves (USD)']

mfg_pmi=im['Manufacturing PMI']; steel=im['Steel Production']
cement_p=im['Cement Production']; elec_p=im['Electricity Production']
coal_p=im['Coal Production']; iip_g=im['IIP Overall Growth']
fert=im['Fertilizer Production']; petro=im['Petroleum Refinery Production']
steel1k=im['Steel Production (1000 Metric Tons)']

tw=tr2['2W Registration']; thw=tr2['3W Registration']
pv_r=tr2['PV Registration']; cv_r=tr2['CV Registration']
tractor=tr2['Tractor Registration']

svc_pmi=se['Service PMI']; svc_exp=se['Service Exports']

nfc=bf['Non-Food Credit']; ms=bf['Money Supply']; liq=bf['Banking Liquidity']
cc=bf['Credit Card Outstanding']; veh_loan=bf['Vehicle Loans']
gold_loan=bf['Loan against gold jewellery']
agri=bf['Credit Deployed (Agri)']; personal=bf['Credit Deployed (Personal)']
ind_c=bf['Credit Deployed (Industries)']; svc_c=bf['Credit Deployed (Services)']

gold_c=cp('Gold'); silver_c=cp('Silver'); copper_c=cp('Copper')
alum_c=cp('Aluminium'); steel_c2=cp('Steel'); lit_c=cp('Lithium')
iron_c=cp('Iron'); freight_c=cp('Freight Rate')

us10y=ty['US-10Y']; us2y=ty['US-2Y']
in10y=ty['India-10Y']; in2y=ty['India-2Y']
cn10y=ty['China-10Y']; jp10y=ty['Japan-10Y']
inr_usd_ty=ty['INR/USD']; jpy_usd=ty['Japanese Yen/USD']; cny_usd=ty['CNY/USD']

def spread(s1,s2):
    d2=dict(s2); return [[k,round(v-d2[k],4)] for k,v in s1 if k in d2]

us_sp=spread(us10y,us2y); in_sp=spread(in10y,in2y)

nifty50=idx['Nifty 50']; nse500=idx['NSE 500']
midcap150=idx['Nifty midcap 150']
sp500=idx['S&P 500']; nasdaq=idx['Nasdaq']
nikkei=idx['Nikkei 225']; shanghai=idx['Shanghai Composite Index']

macro={
    'id':'macro','label':'Macro Economy',
    'title':'India Macro Economy Dashboard',
    'subtitle':'Real macroeconomic data — GST collections, CPI inflation, trade balance, PMI, repo rate, and currency',
    'metrics':[
        {'label':'GST Collections (Rs.B)','value':f'Rs.{gst[-1][1]:.0f}B','change':kch(gst)[0],'up':kch(gst)[1]},
        {'label':'CPI Inflation (%)','value':f'{cpi[-1][1]}%','change':kch(cpi,True,suffix='%')[0],'up':kch(cpi,True)[1]},
        {'label':'Trade Balance (Rs.B)','value':f'Rs.{trade[-1][1]:.0f}B','change':kch(trade,True,suffix='B')[0],'up':kch(trade,True)[1]},
        {'label':'Composite PMI','value':str(pmi_c[-1][1]),'change':kch(pmi_c,True)[0],'up':kch(pmi_c,True)[1]},
        {'label':'Repo Rate (%)','value':f'{rate_i[-1][1]}%','change':kch(rate_i,True,suffix='%')[0],'up':not kch(rate_i,True)[1]},
        {'label':'USD/INR','value':f'Rs.{usd_inr[-1][1]:.2f}','change':kch(usd_inr,True,rnd=2)[0],'up':not kch(usd_inr,True)[1]},
        {'label':'India FX Reserves (USD)','value':f'${fx_res_usd[-1][1]:,.0f} Million','change':kch(fx_res_usd)[0],'up':kch(fx_res_usd)[1]},
        {'label':'India Gold Reserve','value':f'INR {gold_res[-1][1]:,.0f} Cr','change':kch(gold_res)[0],'up':kch(gold_res)[1]},
    ],
    'tableColumns':QC,
    'tableRows':[
        qrow('GST Collections (Rs.B)',gst,lambda x:f'{x:.0f}'),
        qrow('CPI Inflation (%)',cpi),
        qrow('Trade Balance (Rs.B)',trade),
        qrow('Composite PMI',pmi_c),
        qrow('Repo Rate (%)',rate_i),
        qrow('USD/INR',usd_inr,lambda x:f'{x:.2f}'),
        qrow('India Gold Reserve',gold_res),
        qrow('India FX Reserves (USD)',fx_res_usd),
        qrow('India Total FX Reserves',fx_res),
    ],
    'charts':[
        {'title':'GST Collections (Rs. Billion)','unit':'Rs.B','labels':lbl(gst),'series':[{'name':'GST Collections','data':vals(gst),'color':'#6366f1'}]},
        {'title':'Retail Inflation CPI (%)','unit':'%','labels':lbl(cpi),'series':[{'name':'CPI','data':vals(cpi),'color':'#ef4444'}]},
        {'title':'India Trade Balance (Rs.B)','unit':'Rs.B','labels':lbl(trade),'series':[{'name':'Trade Balance','data':vals(trade),'color':'#f59e0b'}]},
        {'title':'Composite PMI','unit':'index','labels':lbl(pmi_c),'series':[{'name':'Composite PMI','data':vals(pmi_c),'color':'#10b981'}]},
        {'title':'India VIX (Volatility Index)','unit':'index','labels':lbl(vix_i),'series':[{'name':'India VIX','data':vals(vix_i),'color':'#8b5cf6'}]},
        {'title':'Repo Rate (%)','unit':'%','labels':lbl(rate_i),'series':[{'name':'Repo Rate','data':vals(rate_i),'color':'#0ea5e9'}]},
        {'title':'USD / INR Exchange Rate','unit':'Rs.','labels':lbl(usd_inr),'series':[{'name':'USD/INR','data':vals(usd_inr),'color':'#f97316'}]},
        {'title':'DXY Dollar Index','unit':'index','labels':lbl(dxy),'series':[{'name':'DXY','data':vals(dxy),'color':'#64748b'}]},
        {'title':'India Gold Reserve (INR Cr)','unit':'INR Cr','labels':lbl(gold_res),'series':[{'name':'Gold Reserve','data':vals(gold_res),'color':'#d97706'}]},
        {'title':'India FX Reserves (USD)','unit':'$M','labels':lbl(fx_res_usd),'series':[{'name':'FX Reserves (USD)','data':vals(fx_res_usd),'color':'#10b981'}]},
    ]
}

industrial={
    'id':'industrial','label':'Industrial & Manufacturing',
    'title':'Industrial & Manufacturing Dashboard',
    'subtitle':"India's core sector — steel, cement, electricity, coal, IIP growth, and manufacturing PMI",
    'metrics':[
        {'label':'Manufacturing PMI','value':str(mfg_pmi[-1][1]),'change':kch(mfg_pmi,True)[0],'up':kch(mfg_pmi,True)[1]},
        {'label':'IIP Overall Growth (%)','value':f'{iip_g[-1][1]}%','change':kch(iip_g,True,suffix='%')[0],'up':kch(iip_g,True)[1]},
        {'label':'Steel Production (MT)','value':str(steel[-1][1]),'change':kch(steel)[0],'up':kch(steel)[1]},
        {'label':'Cement Production','value':str(cement_p[-1][1]),'change':kch(cement_p)[0],'up':kch(cement_p)[1]},
        {'label':'Electricity (BU)','value':str(elec_p[-1][1]),'change':kch(elec_p)[0],'up':kch(elec_p)[1]},
        {'label':'Coal Production','value':str(coal_p[-1][1]),'change':kch(coal_p)[0],'up':kch(coal_p)[1]},
    ],
    'tableColumns':QC,
    'tableRows':[
        qrow('Manufacturing PMI',mfg_pmi),
        qrow('IIP Overall Growth (%)',iip_g),
        qrow('Steel Production (MT)',steel,lambda x:f'{x:.0f}'),
        qrow('Cement Production',cement_p,lambda x:f'{x:.0f}'),
        qrow('Electricity (BU)',elec_p,lambda x:f'{x:.0f}'),
        qrow('Coal Production',coal_p,lambda x:f'{x:.0f}'),
        qrow('Steel Production (Thousand Tons)',steel1k,lambda x:f'{x:,.0f}'),
    ],
    'charts':[
        {'title':'Manufacturing PMI','unit':'index','labels':lbl(mfg_pmi),'series':[{'name':'Mfg PMI','data':vals(mfg_pmi),'color':'#6366f1'}]},
        {'title':'IIP Overall Growth (%)','unit':'%','labels':lbl(iip_g),'series':[{'name':'IIP Overall','data':vals(iip_g),'color':'#10b981'}]},
        {'title':'Steel Production (MT)','unit':'MT','labels':lbl(steel),'series':[{'name':'Steel','data':vals(steel),'color':'#ef4444'}]},
        {'title':'Cement Production (Index)','unit':'index','labels':lbl(cement_p),'series':[{'name':'Cement','data':vals(cement_p),'color':'#8b5cf6'}]},
        {'title':'Electricity Production (BU)','unit':'BU','labels':lbl(elec_p),'series':[{'name':'Electricity','data':vals(elec_p),'color':'#0ea5e9'}]},
        {'title':'Coal Production (Index)','unit':'index','labels':lbl(coal_p),'series':[{'name':'Coal','data':vals(coal_p),'color':'#64748b'}]},
        {'title':'Fertilizer Production (Index)','unit':'index','labels':lbl(fert),'series':[{'name':'Fertilizer','data':vals(fert),'color':'#d97706'}]},
        {'title':'Petroleum Refinery Output (Index)','unit':'index','labels':lbl(petro),'series':[{'name':'Petro Refinery','data':vals(petro),'color':'#be185d'}]},
        {'title':'Steel Production (Thousand Tons)','unit':'Thousand Tons','labels':lbl(steel1k),'series':[{'name':'Steel','data':vals(steel1k),'color':'#ef4444'}]},
    ]
}

transport={
    'id':'transport','label':'Transportation & Automotive',
    'title':'Transportation & Automotive Dashboard',
    'subtitle':'India vehicle registration data — 2W, 3W, Passenger Vehicles, Commercial Vehicles, Tractors',
    'metrics':[
        {'label':'2-Wheeler Registrations','value':f'{tw[-1][1]/100000:.2f}L','change':kch(tw)[0],'up':kch(tw)[1]},
        {'label':'3-Wheeler Registrations','value':f'{thw[-1][1]/1000:.1f}K','change':kch(thw)[0],'up':kch(thw)[1]},
        {'label':'Passenger Vehicles','value':f'{pv_r[-1][1]/1000:.1f}K','change':kch(pv_r)[0],'up':kch(pv_r)[1]},
        {'label':'Commercial Vehicles','value':f'{cv_r[-1][1]/1000:.1f}K','change':kch(cv_r)[0],'up':kch(cv_r)[1]},
        {'label':'Tractor Sales','value':f'{tractor[-1][1]/1000:.1f}K','change':kch(tractor)[0],'up':kch(tractor)[1]},
        {'label':'Total Registered','value':f'{(tw[-1][1]+thw[-1][1]+pv_r[-1][1]+cv_r[-1][1])/100000:.2f}L','change':'+0','up':True},
    ],
    'tableColumns':QC,
    'tableRows':[
        qrow('2W Registration',tw,lambda x:f'{x/1000:.0f}K'),
        qrow('3W Registration',thw,lambda x:f'{x/1000:.1f}K'),
        qrow('Passenger Vehicles',pv_r,lambda x:f'{x/1000:.1f}K'),
        qrow('Commercial Vehicles',cv_r,lambda x:f'{x/1000:.1f}K'),
        qrow('Tractor Registration',tractor,lambda x:f'{x/1000:.1f}K'),
    ],
    'charts':[
        {'title':'2-Wheeler Registrations','unit':'units','labels':lbl(tw),'series':[{'name':'2-Wheeler','data':vals(tw),'color':'#6366f1'}]},
        {'title':'Passenger Vehicle Registrations','unit':'units','labels':lbl(pv_r),'series':[{'name':'PV','data':vals(pv_r),'color':'#10b981'}]},
        {'title':'Commercial Vehicle Registrations','unit':'units','labels':lbl(cv_r),'series':[{'name':'CV','data':vals(cv_r),'color':'#f59e0b'}]},
        {'title':'3-Wheeler Registrations','unit':'units','labels':lbl(thw),'series':[{'name':'3-Wheeler','data':vals(thw),'color':'#0ea5e9'}]},
        {'title':'Tractor Registrations','unit':'units','labels':lbl(tractor),'series':[{'name':'Tractor','data':vals(tractor),'color':'#ef4444'}]},
        {'title':'PV + CV Combined','unit':'units','labels':lbl(pv_r),'series':[
            {'name':'PV','data':vals(pv_r),'color':'#10b981'},
            {'name':'CV','data':vals(cv_r),'color':'#f59e0b'},
        ]},
        {'title':'2W + 3W Combined','unit':'units','labels':lbl(tw),'series':[
            {'name':'2W','data':vals(tw),'color':'#6366f1'},
            {'name':'3W','data':vals(thw),'color':'#8b5cf6'},
        ]},
        {'title':'All Segments Overview','unit':'units','labels':lbl(pv_r),'series':[
            {'name':'PV','data':vals(pv_r),'color':'#10b981'},
            {'name':'CV','data':vals(cv_r),'color':'#f59e0b'},
            {'name':'3W','data':vals(thw),'color':'#8b5cf6'},
        ]},
    ]
}

services={
    'id':'services','label':'Services & Exports',
    'title':'Services & Exports Dashboard',
    'subtitle':"India's services PMI, service exports, PMI comparison, and trade metrics",
    'metrics':[
        {'label':'Services PMI','value':str(svc_pmi[-1][1]),'change':kch(svc_pmi,True)[0],'up':kch(svc_pmi,True)[1]},
        {'label':'Service Exports ($M)','value':f'${svc_exp[-1][1]/1000:.1f}B','change':kch(svc_exp)[0],'up':kch(svc_exp)[1]},
        {'label':'Composite PMI','value':str(pmi_c[-1][1]),'change':kch(pmi_c,True)[0],'up':kch(pmi_c,True)[1]},
        {'label':'Manufacturing PMI','value':str(mfg_pmi[-1][1]),'change':kch(mfg_pmi,True)[0],'up':kch(mfg_pmi,True)[1]},
        {'label':'Trade Balance (Rs.B)','value':f'Rs.{trade[-1][1]:.0f}B','change':kch(trade,True,suffix='B')[0],'up':kch(trade,True)[1]},
        {'label':'USD/INR','value':f'Rs.{usd_inr[-1][1]:.2f}','change':kch(usd_inr,True,rnd=2)[0],'up':not kch(usd_inr,True)[1]},
    ],
    'tableColumns':QC,
    'tableRows':[
        qrow('Services PMI',svc_pmi),
        qrow('Composite PMI',pmi_c),
        qrow('Manufacturing PMI',mfg_pmi),
        qrow('Service Exports ($M)',svc_exp,lambda x:f'{x:.0f}'),
        qrow('Trade Balance (Rs.B)',trade),
        qrow('USD/INR',usd_inr,lambda x:f'{x:.2f}'),
    ],
    'charts':[
        {'title':'Services PMI','unit':'index','labels':lbl(svc_pmi),'series':[{'name':'Services PMI','data':vals(svc_pmi),'color':'#6366f1'}]},
        {'title':'Service Exports ($M)','unit':'$M','labels':lbl(svc_exp),'series':[{'name':'Service Exports','data':vals(svc_exp),'color':'#10b981'}]},
        {'title':'PMI Comparison (Services vs Mfg)','unit':'index','labels':lbl(svc_pmi),'series':[
            {'name':'Services PMI','data':vals(svc_pmi),'color':'#6366f1'},
            {'name':'Manufacturing PMI','data':vals(mfg_pmi),'color':'#f59e0b'},
            {'name':'Composite PMI','data':vals(pmi_c),'color':'#10b981'},
        ]},
        {'title':'India Trade Balance (Rs.B)','unit':'Rs.B','labels':lbl(trade),'series':[{'name':'Trade Balance','data':vals(trade),'color':'#ef4444'}]},
        {'title':'USD/INR Exchange Rate','unit':'Rs.','labels':lbl(usd_inr),'series':[{'name':'USD/INR','data':vals(usd_inr),'color':'#f97316'}]},
        {'title':'DXY Dollar Index','unit':'index','labels':lbl(dxy),'series':[{'name':'DXY','data':vals(dxy),'color':'#64748b'}]},
        {'title':'GST Collections (Rs.B)','unit':'Rs.B','labels':lbl(gst),'series':[{'name':'GST','data':vals(gst),'color':'#8b5cf6'}]},
        {'title':'India VIX','unit':'index','labels':lbl(vix_i),'series':[{'name':'India VIX','data':vals(vix_i),'color':'#0ea5e9'}]},
    ]
}

banking={
    'id':'banking','label':'Banking & Finance',
    'title':'Banking & Finance Dashboard',
    'subtitle':'India banking — non-food credit, money supply, banking liquidity, and sectoral credit deployment',
    'metrics':[
        {'label':'Non-Food Credit','value':f'Rs.{nfc[-1][1]/10000000:.2f}L Cr','change':kch(nfc)[0],'up':kch(nfc)[1]},
        {'label':'Money Supply M1','value':f'Rs.{ms[-1][1]/10000000:.2f}L Cr','change':kch(ms)[0],'up':kch(ms)[1]},
        {'label':'Banking Liquidity (Rs.B)','value':f'Rs.{liq[-1][1]:.0f}B','change':kch(liq,True,suffix='B')[0],'up':liq[-1][1]>0},
        {'label':'Credit Card O/S','value':f'Rs.{cc[-1][1]/100:.0f}Cr','change':kch(cc)[0],'up':kch(cc)[1]},
        {'label':'Vehicle Loans','value':f'Rs.{veh_loan[-1][1]/100:.0f}Cr','change':kch(veh_loan)[0],'up':kch(veh_loan)[1]},
        {'label':'Gold Loans','value':f'Rs.{gold_loan[-1][1]/100:.0f}Cr','change':kch(gold_loan)[0],'up':kch(gold_loan)[1]},
    ],
    'tableColumns':QC,
    'tableRows':[
        qrow('Non-Food Credit (Rs.L Cr)',nfc,lambda x:f'{x/10000000:.2f}'),
        qrow('Money Supply M1 (Rs.L Cr)',ms,lambda x:f'{x/10000000:.2f}'),
        qrow('Banking Liquidity (Rs.B)',liq,lambda x:f'{x:.0f}'),
        qrow('Credit Card O/S (Rs.Cr)',cc,lambda x:f'{x/100:.0f}'),
        qrow('Vehicle Loans (Rs.Cr)',veh_loan,lambda x:f'{x/100:.0f}'),
        qrow('Gold Loans (Rs.Cr)',gold_loan,lambda x:f'{x/100:.0f}'),
    ],
    'charts':[
        {'title':'Non-Food Credit (Rs. Crore)','unit':'Rs.Cr','labels':lbl(nfc),'series':[{'name':'Non-Food Credit','data':vals(nfc),'color':'#6366f1'}]},
        {'title':'Money Supply M1 (Rs. Crore)','unit':'Rs.Cr','labels':lbl(ms),'series':[{'name':'M1','data':vals(ms),'color':'#10b981'}]},
        {'title':'Banking System Liquidity (Rs.B)','unit':'Rs.B','labels':lbl(liq),'series':[{'name':'Liquidity','data':vals(liq),'color':'#ef4444'}]},
        {'title':'Credit Card Outstanding (Rs. Crore)','unit':'Rs.Cr','labels':lbl(cc),'series':[{'name':'CC Outstanding','data':vals(cc),'color':'#f59e0b'}]},
        {'title':'Sectoral Credit Deployment (Rs. Crore)','unit':'Rs.Cr','labels':lbl(agri),'series':[
            {'name':'Agriculture','data':vals(agri),'color':'#10b981'},
            {'name':'Industry','data':vals(ind_c),'color':'#6366f1'},
            {'name':'Services','data':vals(svc_c),'color':'#f59e0b'},
        ]},
        {'title':'Vehicle Loans (Rs. Crore)','unit':'Rs.Cr','labels':lbl(veh_loan),'series':[{'name':'Vehicle Loans','data':vals(veh_loan),'color':'#8b5cf6'}]},
        {'title':'Gold Jewellery Loans (Rs. Crore)','unit':'Rs.Cr','labels':lbl(gold_loan),'series':[{'name':'Gold Loans','data':vals(gold_loan),'color':'#d97706'}]},
        {'title':'Personal Credit (Rs. Crore)','unit':'Rs.Cr','labels':lbl(personal),'series':[{'name':'Personal Credit','data':vals(personal),'color':'#be185d'}]},
    ]
}

commodities={
    'id':'commodities','label':'Commodity Prices',
    'title':'Commodity Prices Dashboard',
    'subtitle':'Real commodity prices — gold, silver, copper, aluminium, steel, lithium, iron ore, and freight',
    'metrics':[
        {'label':'Gold (Rs./100g)','value':f'Rs.{gold_c[-1][1]/10:.0f}','change':kch(gold_c)[0],'up':kch(gold_c)[1]},
        {'label':'Silver (Rs.)','value':f'Rs.{silver_c[-1][1]/1000:.0f}K','change':kch(silver_c)[0],'up':kch(silver_c)[1]},
        {'label':'Copper (Rs./MT)','value':f'Rs.{copper_c[-1][1]/1000:.1f}K','change':kch(copper_c)[0],'up':kch(copper_c)[1]},
        {'label':'Aluminium (Rs./MT)','value':f'Rs.{alum_c[-1][1]/1000:.1f}K','change':kch(alum_c)[0],'up':kch(alum_c)[1]},
        {'label':'Steel (Rs./MT)','value':f'Rs.{steel_c2[-1][1]/1000:.1f}K','change':kch(steel_c2)[0],'up':kch(steel_c2)[1]},
        {'label':'Lithium (Rs./MT)','value':f'Rs.{lit_c[-1][1]/1000:.0f}K','change':kch(lit_c)[0],'up':kch(lit_c)[1]},
    ],
    'tableColumns':QC,
    'tableRows':[
        qrow('Gold (Rs.)',gold_c,lambda x:f'{x:.0f}'),
        qrow('Silver (Rs.)',silver_c,lambda x:f'{x:.0f}'),
        qrow('Copper (Rs./MT)',copper_c,lambda x:f'{x:.0f}'),
        qrow('Aluminium (Rs./MT)',alum_c,lambda x:f'{x:.0f}'),
        qrow('Steel (Rs./MT)',steel_c2,lambda x:f'{x:.0f}'),
        qrow('Lithium (Rs./MT)',lit_c,lambda x:f'{x:.0f}'),
    ],
    'charts':[
        {'title':'Gold Price (Rs.)','unit':'Rs.','labels':lbl(gold_c),'series':[{'name':'Gold','data':vals(gold_c),'color':'#d97706'}]},
        {'title':'Silver Price (Rs.)','unit':'Rs.','labels':lbl(silver_c),'series':[{'name':'Silver','data':vals(silver_c),'color':'#64748b'}]},
        {'title':'Copper Price (Rs./MT)','unit':'Rs./MT','labels':lbl(copper_c),'series':[{'name':'Copper','data':vals(copper_c),'color':'#b45309'}]},
        {'title':'Aluminium Price (Rs./MT)','unit':'Rs./MT','labels':lbl(alum_c),'series':[{'name':'Aluminium','data':vals(alum_c),'color':'#6366f1'}]},
        {'title':'Steel Price (Rs./MT)','unit':'Rs./MT','labels':lbl(steel_c2),'series':[{'name':'Steel','data':vals(steel_c2),'color':'#ef4444'}]},
        {'title':'Lithium Price (Rs./MT)','unit':'Rs./MT','labels':lbl(lit_c),'series':[{'name':'Lithium','data':vals(lit_c),'color':'#10b981'}]},
        {'title':'Iron Ore Price ($/MT)','unit':'$/MT','labels':lbl(iron_c),'series':[{'name':'Iron Ore','data':vals(iron_c),'color':'#8b5cf6'}]},
        {'title':'Baltic Dry Index (Freight)','unit':'index','labels':lbl(freight_c),'series':[{'name':'BDI','data':vals(freight_c),'color':'#0ea5e9'}]},
    ]
}

treasury={
    'id':'treasury','label':'Treasury Yield',
    'title':'Global Treasury Yield Dashboard',
    'subtitle':'Sovereign bond yields — India, US, China & Japan — with yield spreads and currency rates',
    'metrics':[
        {'label':'India 10Y (%)','value':f'{in10y[-1][1]}%','change':kch(in10y,True,suffix='%')[0],'up':not kch(in10y,True)[1]},
        {'label':'US 10Y (%)','value':f'{us10y[-1][1]}%','change':kch(us10y,True,suffix='%')[0],'up':not kch(us10y,True)[1]},
        {'label':'China 10Y (%)','value':f'{cn10y[-1][1]}%','change':kch(cn10y,True,suffix='%')[0],'up':not kch(cn10y,True)[1]},
        {'label':'Japan 10Y (%)','value':f'{jp10y[-1][1]}%','change':kch(jp10y,True,suffix='%')[0],'up':not kch(jp10y,True)[1]},
        {'label':'US 10Y-2Y Spread','value':f'{us_sp[-1][1]}%','change':kch(us_sp,True,suffix='%')[0],'up':kch(us_sp,True)[1]},
        {'label':'USD/INR','value':f'Rs.{inr_usd_ty[-1][1]:.2f}','change':kch(inr_usd_ty,True,rnd=2)[0],'up':not kch(inr_usd_ty,True)[1]},
    ],
    'tableColumns':QC,
    'tableRows':[
        qrow('India 10Y (%)',in10y),
        qrow('India 2Y (%)',in2y),
        qrow('US 10Y (%)',us10y),
        qrow('US 2Y (%)',us2y),
        qrow('China 10Y (%)',cn10y),
        qrow('Japan 10Y (%)',jp10y),
    ],
    'charts':[
        {'title':'India Treasury Yields (%)','unit':'%','labels':lbl(in10y),'series':[
            {'name':'India 10Y','data':vals(in10y),'color':'#6366f1'},
            {'name':'India 2Y','data':vals(in2y),'color':'#a5b4fc'},
        ]},
        {'title':'US Treasury Yields (%)','unit':'%','labels':lbl(us10y),'series':[
            {'name':'US 10Y','data':vals(us10y),'color':'#ef4444'},
            {'name':'US 2Y','data':vals(us2y),'color':'#fca5a5'},
        ]},
        {'title':'Global 10Y Yields (%)','unit':'%','labels':lbl(in10y),'series':[
            {'name':'India','data':vals(in10y),'color':'#6366f1'},
            {'name':'US','data':vals(us10y),'color':'#ef4444'},
            {'name':'China','data':vals(cn10y),'color':'#f59e0b'},
            {'name':'Japan','data':vals(jp10y),'color':'#10b981'},
        ]},
        {'title':'US Yield Curve Spread 10Y-2Y (%)','unit':'%','labels':lbl(us_sp),'series':[{'name':'10Y-2Y','data':vals(us_sp),'color':'#8b5cf6'}]},
        {'title':'India Yield Spread 10Y-2Y (%)','unit':'%','labels':lbl(in_sp),'series':[{'name':'India Spread','data':vals(in_sp),'color':'#0ea5e9'}]},
        {'title':'USD/INR Rate','unit':'Rs.','labels':lbl(inr_usd_ty),'series':[{'name':'USD/INR','data':vals(inr_usd_ty),'color':'#f97316'}]},
        {'title':'JPY/USD Rate','unit':'JPY','labels':lbl(jpy_usd),'series':[{'name':'JPY/USD','data':vals(jpy_usd),'color':'#d97706'}]},
        {'title':'CNY/USD Rate','unit':'CNY','labels':lbl(cny_usd),'series':[{'name':'CNY/USD','data':vals(cny_usd),'color':'#64748b'}]},
    ]
}

equity={
    'id':'equity','label':'Equity / Indices',
    'title':'Global Equity & Indices Dashboard',
    'subtitle':'Nifty 50, NSE 500, S&P 500, Nasdaq, Nikkei 225, Shanghai Composite, and EURO STOXX 50',
    'metrics':[
        {'label':'Nifty 50','value':f'{nifty50[-1][1]:.0f}','change':kch(nifty50)[0],'up':kch(nifty50)[1]},
        {'label':'NSE 500','value':f'{nse500[-1][1]:.0f}','change':kch(nse500)[0],'up':kch(nse500)[1]},
        {'label':'S&P 500','value':f'{sp500[-1][1]:.0f}','change':kch(sp500)[0],'up':kch(sp500)[1]},
        {'label':'Nasdaq','value':f'{nasdaq[-1][1]:.0f}','change':kch(nasdaq)[0],'up':kch(nasdaq)[1]},
        {'label':'Nikkei 225','value':f'{nikkei[-1][1]:.0f}','change':kch(nikkei)[0],'up':kch(nikkei)[1]},
        {'label':'Shanghai Comp.','value':f'{shanghai[-1][1]:.0f}','change':kch(shanghai)[0],'up':kch(shanghai)[1]},
    ],
    'tableColumns':QC,
    'tableRows':[
        qrow('Nifty 50',nifty50,lambda x:f'{x:.0f}'),
        qrow('NSE 500',nse500,lambda x:f'{x:.0f}'),
        qrow('S&P 500',sp500,lambda x:f'{x:.0f}'),
        qrow('Nasdaq',nasdaq,lambda x:f'{x:.0f}'),
        qrow('Nikkei 225',nikkei,lambda x:f'{x:.0f}'),
        qrow('Shanghai Comp.',shanghai,lambda x:f'{x:.0f}'),
    ],
    'charts':[
        {'title':'Nifty 50 Index','unit':'pts','labels':lbl(nifty50),'series':[{'name':'Nifty 50','data':vals(nifty50),'color':'#6366f1'}]},
        {'title':'NSE 500 Index','unit':'pts','labels':lbl(nse500),'series':[{'name':'NSE 500','data':vals(nse500),'color':'#8b5cf6'}]},
        {'title':'Nifty Midcap 150','unit':'pts','labels':lbl(midcap150),'series':[{'name':'Midcap 150','data':vals(midcap150),'color':'#0ea5e9'}]},
        {'title':'Indian Indices Comparison','unit':'pts','labels':lbl(nifty50),'series':[
            {'name':'Nifty 50','data':vals(nifty50),'color':'#6366f1'},
            {'name':'NSE 500','data':vals(nse500),'color':'#8b5cf6'},
        ]},
        {'title':'S&P 500 Index','unit':'pts','labels':lbl(sp500),'series':[{'name':'S&P 500','data':vals(sp500),'color':'#ef4444'}]},
        {'title':'Nasdaq Composite','unit':'pts','labels':lbl(nasdaq),'series':[{'name':'Nasdaq','data':vals(nasdaq),'color':'#f59e0b'}]},
        {'title':'Nikkei 225','unit':'pts','labels':lbl(nikkei),'series':[{'name':'Nikkei 225','data':vals(nikkei),'color':'#10b981'}]},
        {'title':'Global Indices Comparison','unit':'pts','labels':lbl(sp500),'series':[
            {'name':'S&P 500','data':vals(sp500),'color':'#ef4444'},
            {'name':'Nasdaq','data':vals(nasdaq),'color':'#f59e0b'},
        ]},
    ]
}

all_ind=[macro,industrial,transport,services,banking,commodities,treasury,equity]
js='// Generated from real Excel data — Economic Quarterly data Final.xlsx\n'
js+='// All values are actual, not simulated. Last update: Feb 2026\n\n'
js+=f'const INDUSTRIES = {json.dumps(all_ind, indent=2)};\n\n'
js+='export const INDUSTRY_MAP = Object.fromEntries(INDUSTRIES.map((i) => [i.id, i]));\n\nexport default INDUSTRIES;\n'

with open(OUT,'w',encoding='utf-8') as f:
    f.write(js)
print(f'Written {os.path.getsize(OUT)} bytes')
print('IDs:', [i['id'] for i in all_ind])
